"""
Exportadores a Excel y PDF - AlmacénPro v2.0
Utilidades para exportar datos a diferentes formatos
"""

import os
import logging
from datetime import datetime
from typing import Dict, List, Any, Optional
from io import BytesIO

# Formatters internos
from .formatters import NumberFormatter, DateFormatter, TextFormatter

logger = logging.getLogger(__name__)

class ExcelExporter:
    """Exportador a archivos Excel"""
    
    def __init__(self):
        self.openpyxl_available = False
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
            self.openpyxl = openpyxl
            self.Font = Font
            self.Alignment = Alignment
            self.PatternFill = PatternFill
            self.Border = Border
            self.Side = Side
            self.openpyxl_available = True
        except ImportError:
            logger.warning("openpyxl no disponible. Exportación a Excel deshabilitada.")
    
    def export_data(self, data: List[Dict], filename: str, 
                   sheet_name: str = "Datos", 
                   headers: Optional[List[str]] = None,
                   title: Optional[str] = None) -> bool:
        """Exportar datos a Excel"""
        if not self.openpyxl_available:
            logger.error("openpyxl no disponible")
            return False
        
        try:
            # Crear workbook y hoja
            wb = self.openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            current_row = 1
            
            # Añadir título si se proporciona
            if title:
                ws.merge_cells(f'A1:Z1')
                title_cell = ws['A1']
                title_cell.value = title
                title_cell.font = self.Font(size=16, bold=True)
                title_cell.alignment = self.Alignment(horizontal='center')
                current_row = 3
            
            if not data:
                ws.cell(row=current_row, column=1, value="No hay datos para exportar")
                wb.save(filename)
                return True
            
            # Determinar headers
            if headers is None:
                headers = list(data[0].keys()) if data else []
            
            # Escribir headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = self.Font(bold=True)
                cell.fill = self.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = self.Alignment(horizontal='center')
            
            current_row += 1
            
            # Escribir datos
            for row_data in data:
                for col, header in enumerate(headers, 1):
                    value = row_data.get(header, "")
                    
                    # Formatear valor según tipo
                    if isinstance(value, (int, float)):
                        if any(keyword in str(header).lower() for keyword in ['precio', 'monto', 'total', 'importe']):
                            formatted_value = NumberFormatter.format_currency(value)
                        else:
                            formatted_value = NumberFormatter.format_number(value)
                    elif isinstance(value, datetime):
                        formatted_value = DateFormatter.format_datetime(value)
                    else:
                        formatted_value = str(value) if value is not None else ""
                    
                    ws.cell(row=current_row, column=col, value=formatted_value)
                
                current_row += 1
            
            # Autoajustar columnas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Guardar archivo
            wb.save(filename)
            logger.info(f"Datos exportados a Excel: {filename}")
            return True
            
        except Exception as e:
            logger.error(f"Error exportando a Excel: {e}")
            return False
    
    def export_customer_report(self, customers: List[Dict], filename: str) -> bool:
        """Exportar reporte de clientes a Excel"""
        if not self.openpyxl_available:
            return False
        
        try:
            wb = self.openpyxl.Workbook()
            ws = wb.active
            ws.title = "Clientes"
            
            # Título
            ws.merge_cells('A1:H1')
            title_cell = ws['A1']
            title_cell.value = f"Reporte de Clientes - {DateFormatter.format_datetime(datetime.now())}"
            title_cell.font = self.Font(size=14, bold=True)
            title_cell.alignment = self.Alignment(horizontal='center')
            
            # Headers
            headers = ["ID", "Nombre", "Email", "Teléfono", "Categoría", "Límite Crédito", "Saldo", "Estado"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = self.Font(bold=True)
                cell.fill = self.PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
                cell.font = self.Font(color="FFFFFF", bold=True)
            
            # Datos
            row = 4
            for customer in customers:
                ws.cell(row=row, column=1, value=customer.get('id', ''))
                ws.cell(row=row, column=2, value=f"{customer.get('nombre', '')} {customer.get('apellido', '')}")
                ws.cell(row=row, column=3, value=customer.get('email', ''))
                ws.cell(row=row, column=4, value=customer.get('telefono', ''))
                ws.cell(row=row, column=5, value=customer.get('categoria_cliente', ''))
                ws.cell(row=row, column=6, value=NumberFormatter.format_currency(customer.get('limite_credito', 0)))
                ws.cell(row=row, column=7, value=NumberFormatter.format_currency(customer.get('saldo_cuenta_corriente', 0)))
                ws.cell(row=row, column=8, value="Activo" if customer.get('activo', True) else "Inactivo")
                row += 1
            
            # Autoajustar columnas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filename)
            return True
            
        except Exception as e:
            logger.error(f"Error exportando reporte de clientes: {e}")
            return False


class PDFExporter:
    """Exportador a archivos PDF"""
    
    def __init__(self):
        self.reportlab_available = False
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.lib import colors
            
            self.letter = letter
            self.A4 = A4
            self.SimpleDocTemplate = SimpleDocTemplate
            self.Table = Table
            self.TableStyle = TableStyle
            self.Paragraph = Paragraph
            self.Spacer = Spacer
            self.getSampleStyleSheet = getSampleStyleSheet
            self.ParagraphStyle = ParagraphStyle
            self.inch = inch
            self.colors = colors
            self.reportlab_available = True
            
        except ImportError:
            logger.warning("reportlab no disponible. Exportación a PDF deshabilitada.")
    
    def export_data(self, data: List[Dict], filename: str,
                   title: str = "Reporte", 
                   headers: Optional[List[str]] = None,
                   company_name: str = "AlmacénPro") -> bool:
        """Exportar datos a PDF"""
        if not self.reportlab_available:
            logger.error("reportlab no disponible")
            return False
        
        try:
            # Crear documento
            doc = self.SimpleDocTemplate(filename, pagesize=self.A4)
            story = []
            
            # Estilos
            styles = self.getSampleStyleSheet()
            title_style = self.ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=1  # Centrado
            )
            
            # Título del documento
            story.append(self.Paragraph(company_name, title_style))
            story.append(self.Paragraph(title, styles['Heading2']))
            story.append(self.Paragraph(f"Generado: {DateFormatter.format_datetime(datetime.now())}", styles['Normal']))
            story.append(self.Spacer(1, 20))
            
            if not data:
                story.append(self.Paragraph("No hay datos para mostrar", styles['Normal']))
                doc.build(story)
                return True
            
            # Determinar headers
            if headers is None:
                headers = list(data[0].keys()) if data else []
            
            # Preparar datos para tabla
            table_data = [headers]  # Primera fila con headers
            
            for row_data in data:
                formatted_row = []
                for header in headers:
                    value = row_data.get(header, "")
                    
                    # Formatear valor
                    if isinstance(value, (int, float)):
                        if any(keyword in str(header).lower() for keyword in ['precio', 'monto', 'total', 'importe']):
                            formatted_value = NumberFormatter.format_currency(value)
                        else:
                            formatted_value = NumberFormatter.format_number(value)
                    elif isinstance(value, datetime):
                        formatted_value = DateFormatter.format_datetime(value)
                    else:
                        formatted_value = str(value) if value is not None else ""
                    
                    # Truncar si es muy largo
                    formatted_value = TextFormatter.truncate(formatted_value, 25)
                    formatted_row.append(formatted_value)
                
                table_data.append(formatted_row)
            
            # Crear tabla
            table = self.Table(table_data)
            
            # Estilo de tabla
            table.setStyle(self.TableStyle([
                # Header
                ('BACKGROUND', (0, 0), (-1, 0), self.colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), self.colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                
                # Datos
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [self.colors.beige, self.colors.white]),
                ('GRID', (0, 0), (-1, -1), 1, self.colors.black)
            ]))
            
            story.append(table)
            
            # Construir documento
            doc.build(story)
            logger.info(f"Datos exportados a PDF: {filename}")
            return True
            
        except Exception as e:
            logger.error(f"Error exportando a PDF: {e}")
            return False
    
    def export_invoice(self, sale_data: Dict, filename: str) -> bool:
        """Exportar factura/comprobante a PDF"""
        if not self.reportlab_available:
            return False
        
        try:
            doc = self.SimpleDocTemplate(filename, pagesize=self.letter)
            story = []
            
            styles = self.getSampleStyleSheet()
            
            # Header de factura
            story.append(self.Paragraph("ALMACÉN PRO", styles['Title']))
            story.append(self.Paragraph("Sistema de Gestión Empresarial", styles['Normal']))
            story.append(self.Spacer(1, 20))
            
            # Información de la venta
            sale_info = f"""
            <b>Factura N°:</b> {sale_data.get('numero_factura', 'N/A')}<br/>
            <b>Fecha:</b> {DateFormatter.format_datetime(sale_data.get('fecha_venta', datetime.now()))}<br/>
            <b>Cliente:</b> {sale_data.get('cliente_nombre', 'Consumidor Final')}<br/>
            """
            
            story.append(self.Paragraph(sale_info, styles['Normal']))
            story.append(self.Spacer(1, 20))
            
            # Tabla de productos
            items = sale_data.get('items', [])
            if items:
                table_data = [['Producto', 'Cantidad', 'Precio Unit.', 'Subtotal']]
                
                for item in items:
                    table_data.append([
                        item.get('producto_nombre', ''),
                        str(item.get('cantidad', 0)),
                        NumberFormatter.format_currency(item.get('precio_unitario', 0)),
                        NumberFormatter.format_currency(item.get('subtotal', 0))
                    ])
                
                # Totales
                table_data.append(['', '', 'Subtotal:', NumberFormatter.format_currency(sale_data.get('subtotal', 0))])
                table_data.append(['', '', 'Impuestos:', NumberFormatter.format_currency(sale_data.get('impuestos', 0))])
                table_data.append(['', '', 'TOTAL:', NumberFormatter.format_currency(sale_data.get('total', 0))])
                
                table = self.Table(table_data)
                table.setStyle(self.TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), self.colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), self.colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 1, self.colors.black),
                    ('BACKGROUND', (0, -3), (-1, -1), self.colors.lightgrey),
                    ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ]))
                
                story.append(table)
            
            doc.build(story)
            return True
            
        except Exception as e:
            logger.error(f"Error exportando factura: {e}")
            return False


class CSVExporter:
    """Exportador a archivos CSV"""
    
    @staticmethod
    def export_data(data: List[Dict], filename: str, 
                   headers: Optional[List[str]] = None,
                   delimiter: str = ',') -> bool:
        """Exportar datos a CSV"""
        try:
            import csv
            
            if not data:
                with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                    csvfile.write("No hay datos para exportar")
                return True
            
            # Determinar headers
            if headers is None:
                headers = list(data[0].keys())
            
            with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=headers, delimiter=delimiter)
                
                # Escribir headers
                writer.writeheader()
                
                # Escribir datos
                for row in data:
                    formatted_row = {}
                    for header in headers:
                        value = row.get(header, "")
                        
                        # Formatear valor
                        if isinstance(value, (int, float)):
                            if any(keyword in str(header).lower() for keyword in ['precio', 'monto', 'total', 'importe']):
                                formatted_value = NumberFormatter.format_currency(value)
                            else:
                                formatted_value = NumberFormatter.format_number(value)
                        elif isinstance(value, datetime):
                            formatted_value = DateFormatter.format_datetime(value)
                        else:
                            formatted_value = str(value) if value is not None else ""
                        
                        formatted_row[header] = formatted_value
                    
                    writer.writerow(formatted_row)
            
            logger.info(f"Datos exportados a CSV: {filename}")
            return True
            
        except Exception as e:
            logger.error(f"Error exportando a CSV: {e}")
            return False


class FileExporter:
    """Exportador principal que maneja múltiples formatos"""
    
    def __init__(self):
        self.excel_exporter = ExcelExporter()
        self.pdf_exporter = PDFExporter()
        self.csv_exporter = CSVExporter()
    
    def export(self, data: List[Dict], filename: str, format_type: str = "excel",
               title: str = "Reporte", headers: Optional[List[str]] = None, **kwargs) -> bool:
        """Exportar datos al formato especificado"""
        try:
            # Determinar formato por extensión si no se especifica
            if format_type == "auto":
                ext = os.path.splitext(filename)[1].lower()
                if ext in ['.xlsx', '.xls']:
                    format_type = "excel"
                elif ext == '.pdf':
                    format_type = "pdf"
                elif ext == '.csv':
                    format_type = "csv"
                else:
                    format_type = "excel"  # Default
            
            # Exportar según formato
            if format_type.lower() == "excel":
                return self.excel_exporter.export_data(data, filename, title, headers)
            elif format_type.lower() == "pdf":
                return self.pdf_exporter.export_data(data, filename, title, headers, **kwargs)
            elif format_type.lower() == "csv":
                return self.csv_exporter.export_data(data, filename, headers, **kwargs)
            else:
                logger.error(f"Formato no soportado: {format_type}")
                return False
                
        except Exception as e:
            logger.error(f"Error en exportación: {e}")
            return False
    
    def export_customers(self, customers: List[Dict], filename: str, format_type: str = "excel") -> bool:
        """Exportar lista de clientes"""
        headers = ["id", "nombre", "apellido", "email", "telefono", "categoria_cliente", 
                  "limite_credito", "saldo_cuenta_corriente", "activo"]
        
        return self.export(customers, filename, format_type, "Reporte de Clientes", headers)
    
    def export_sales(self, sales: List[Dict], filename: str, format_type: str = "excel") -> bool:
        """Exportar reporte de ventas"""
        headers = ["id", "numero_factura", "fecha_venta", "cliente_nombre", 
                  "subtotal", "impuestos", "total", "estado"]
        
        return self.export(sales, filename, format_type, "Reporte de Ventas", headers)
    
    def export_products(self, products: List[Dict], filename: str, format_type: str = "excel") -> bool:
        """Exportar catálogo de productos"""
        headers = ["codigo_barras", "nombre", "descripcion", "categoria_nombre", 
                  "precio_venta", "precio_compra", "stock_actual", "activo"]
        
        return self.export(products, filename, format_type, "Catálogo de Productos", headers)
    
    def is_format_available(self, format_type: str) -> bool:
        """Verificar si un formato está disponible"""
        if format_type.lower() == "excel":
            return self.excel_exporter.openpyxl_available
        elif format_type.lower() == "pdf":
            return self.pdf_exporter.reportlab_available
        elif format_type.lower() == "csv":
            return True  # CSV siempre disponible
        else:
            return False
    
    def get_available_formats(self) -> List[str]:
        """Obtener lista de formatos disponibles"""
        formats = ["csv"]  # CSV siempre disponible
        
        if self.excel_exporter.openpyxl_available:
            formats.append("excel")
        
        if self.pdf_exporter.reportlab_available:
            formats.append("pdf")
        
        return formats


# Función de conveniencia
def export_data(data: List[Dict], filename: str, format_type: str = "excel", 
               title: str = "Reporte", headers: Optional[List[str]] = None) -> bool:
    """Función de conveniencia para exportar datos"""
    exporter = FileExporter()
    return exporter.export(data, filename, format_type, title, headers)

def export_to_excel(data: List[Dict], filename: str, title: str = "Reporte") -> bool:
    """Función de conveniencia para exportar a Excel"""
    exporter = ExcelExporter()
    return exporter.export_data(data, filename, title=title)

def export_to_pdf(data: List[Dict], filename: str, title: str = "Reporte") -> bool:
    """Función de conveniencia para exportar a PDF"""
    exporter = PDFExporter()
    return exporter.export_data(data, filename, title=title)

def export_to_csv(data: List[Dict], filename: str) -> bool:
    """Función de conveniencia para exportar a CSV"""
    return CSVExporter.export_data(data, filename)